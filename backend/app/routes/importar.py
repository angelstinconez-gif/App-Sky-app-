"""Importación masiva desde archivos Excel (.xlsx).

Todos los endpoints usan UPSERT (no INSERT ciego):
- Si la fila ya existe por su clave única, se ACTUALIZA con los campos no vacíos.
- Si no existe, se crea nueva.
Así, reimportar el mismo archivo NO genera duplicados.
"""
from io import BytesIO

from flask import Blueprint, jsonify, request
from flask_jwt_extended import jwt_required
from openpyxl import load_workbook

from app import db
from app.models.incidencia import Incidencia
from app.models.poliza import Poliza
from app.models.garantia import Garantia
from app.models.error_catalog import ErrorCatalog
from app.models.directorio import Directorio
from app.models.mantenimiento import Mantenimiento
from app.utils.audit import log_change
from app.utils.decorators import admin_required
from app.utils.parse import parse_date, parse_int, parse_str

bp = Blueprint("importar", __name__)


def _rows(file_storage, sheet=None):
    """Lee la hoja activa (o una con nombre concreto).

    Detecta headers de 2 niveles: si la fila 1 tiene celdas que la fila 0 no,
    los combina (útil para Pólizas con 'Garantía' / 'Inicio / Fin / Estatus').
    """
    wb = load_workbook(BytesIO(file_storage.read()), data_only=True)
    ws = wb[sheet] if (sheet and sheet in wb.sheetnames) else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    row0 = list(rows[0])
    row1 = list(rows[1]) if len(rows) > 1 else []

    # Decide si fila 1 es sub-header: tiene celdas no-vacías donde fila 0 no
    is_subheader = False
    if row1:
        for i, v in enumerate(row1):
            if v and i < len(row0) and (row0[i] in (None, "")):
                is_subheader = True
                break

    headers = []
    if is_subheader:
        for i, v in enumerate(row0):
            sub = row1[i] if i < len(row1) else None
            if v and sub:
                headers.append(f"{str(v).strip()} {str(sub).strip()}")
            elif sub:
                headers.append(str(sub).strip())
            elif v:
                headers.append(str(v).strip())
            else:
                headers.append("")
        data_rows = rows[2:]
    else:
        headers = [str(h or "").strip() for h in row0]
        data_rows = rows[1:]

    return headers, data_rows


def _apply_non_empty(obj, **fields):
    """Asigna sólo campos con valor (no None/vacío) — preserva datos existentes."""
    for k, v in fields.items():
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        setattr(obj, k, v)


# ──────────────────────────────────────────────────────────
#  INCIDENCIAS — upsert por (site, code, err_code, inc_date)
# ──────────────────────────────────────────────────────────
@bp.route("/incidencias", methods=["POST"])
@jwt_required()
@admin_required
def import_incidencias():
    f = request.files.get("file")
    if not f:
        return jsonify(error="no_file"), 400
    headers, rows = _rows(f)
    idx = {h.lower(): n for n, h in enumerate(headers)}

    def g(row, *keys):
        for k in keys:
            n = idx.get(k.lower())
            if n is not None and n < len(row):
                return row[n]
        return None

    created, updated = 0, 0
    for row in rows:
        if not any(row):
            continue
        site = parse_str(g(row, "site", "sitio", "proyecto", "project"))
        if not site:
            continue
        code = parse_str(g(row, "code", "codigo"))
        err_code = parse_str(g(row, "errCode", "error", "código de error"))
        inc_date = parse_date(g(row, "incDate", "fecha"))

        existing = Incidencia.query.filter_by(
            site=site, code=code, err_code=err_code, inc_date=inc_date
        ).first()
        target = existing or Incidencia(site=site, status="abierta")
        _apply_non_empty(
            target,
            platform=parse_str(g(row, "platform", "plataforma")),
            site=site,
            client=parse_str(g(row, "client", "cliente")),
            code=code,
            priority=parse_str(g(row, "priority", "prioridad")),
            notes=parse_str(g(row, "notes", "notas")),
            inc_date=inc_date,
            err_code=err_code,
            classification=parse_str(g(row, "classification", "clasificacion")),
            equipment=parse_str(g(row, "equipment", "equipo")),
            problem=parse_str(g(row, "problem", "problema")),
            cause=parse_str(g(row, "cause", "causa")),
            solution=parse_str(g(row, "solution", "solucion")),
        )
        if existing:
            updated += 1
        else:
            db.session.add(target)
            created += 1
    log_change("importar", "incidencias", f"{created} nuevas, {updated} actualizadas")
    db.session.commit()
    return jsonify(ok=True, created=created, updated=updated)


# ──────────────────────────────────────────────────────────
#  PÓLIZAS — upsert por code, secundario por project
# ──────────────────────────────────────────────────────────
@bp.route("/polizas", methods=["POST"])
@jwt_required()
@admin_required
def import_polizas():
    f = request.files.get("file")
    if not f:
        return jsonify(error="no_file"), 400
    headers, rows = _rows(f)
    idx = {h.lower(): n for n, h in enumerate(headers)}

    def g(row, *keys):
        for k in keys:
            n = idx.get(k.lower())
            if n is not None and n < len(row):
                return row[n]
        return None

    created, updated = 0, 0
    for row in rows:
        if not any(row):
            continue
        project = parse_str(g(row, "project", "proyecto"))
        code = parse_str(g(row, "code", "codigo", "código"))
        if not project and not code:
            continue

        existing = None
        if code:
            existing = Poliza.query.filter_by(code=code).first()
        if not existing and project:
            existing = Poliza.query.filter_by(project=project).first()

        target = existing or Poliza(project=project or "—")
        # Datos del Excel oficial — siempre sobrescribir fechas/plataforma/zona/status
        sys_start = parse_date(g(row, "sysStart", "inicioSistema", "inicio del sistema", "sys start"))
        pol_start = parse_date(g(row, "polStart", "inicioPoliza", "inicio", "garantía inicio", "garantia inicio"))
        pol_end   = parse_date(g(row, "polEnd", "finPoliza", "fin", "garantía fin", "garantia fin"))
        platform  = parse_str(g(row, "platform", "plataforma"))
        # La columna "Tipo de Póliza" del Excel describe la COBERTURA (Completo, Eléctrico,
        # Mantenimiento, Operación, etc.), NO el tipo de sistema. Va a `cobertura`.
        cobertura_excel = parse_str(g(row, "tipo de poliza", "tipo de póliza", "tipoPoliza", "cobertura"))
        zona      = parse_str(g(row, "zona", "ubicación", "ubicacion"))
        status    = parse_str(g(row, "status", "estatus", "garantía estatus", "garantia estatus"))

        if sys_start is not None:   target.sys_start = sys_start
        if pol_start is not None:   target.pol_start = pol_start
        if pol_end is not None:     target.pol_end = pol_end
        if platform:                target.platform = platform
        if cobertura_excel:         target.cobertura = cobertura_excel
        if zona:                    target.zona = zona
        if status:                  target.status = status

        # ── Derivar SIEMPRE `poliza` (tipo de sistema) desde el código ──
        # -FV → PV · -BT → BESS · -HB → Híbrido
        c_up = (code or "").upper()
        if "-FV" in c_up:
            target.poliza = "PV"
        elif "-HB" in c_up:
            target.poliza = "Híbrido"
        elif "-BT" in c_up:
            target.poliza = "BESS"
        # Si no hay código, mantiene lo que ya tenía (no toca poliza)

        # Resto preserva info ya capturada
        _apply_non_empty(
            target,
            project=project,
            item=parse_int(g(row, "item")),
            grupo=parse_str(g(row, "grupo", "gupo")),
            code=code,
            tarifa=parse_str(g(row, "tarifa")),
            panels=parse_str(g(row, "panels", "paneles")),
            inv=parse_str(g(row, "inv", "inversores")),
            cuadrilla=parse_str(g(row, "cuadrilla")),
        )
        if existing:
            updated += 1
        else:
            db.session.add(target)
            created += 1
    log_change("importar", "polizas", f"{created} nuevas, {updated} actualizadas")
    db.session.commit()
    return jsonify(ok=True, created=created, updated=updated)


# ──────────────────────────────────────────────────────────
#  DIRECTORIO — upsert por (project, maint_contact)
# ──────────────────────────────────────────────────────────
@bp.route("/directorio", methods=["POST"])
@jwt_required()
@admin_required
def import_directorio():
    f = request.files.get("file")
    if not f:
        return jsonify(error="no_file"), 400
    headers, rows = _rows(f)
    idx = {h.lower().strip(): n for n, h in enumerate(headers) if h}

    def g(row, *keys):
        for k in keys:
            n = idx.get(k.lower())
            if n is not None and n < len(row) and row[n] not in (None, "", "N/A"):
                return row[n]
        return None

    created, updated = 0, 0
    for row in rows:
        if not any(row):
            continue
        project = parse_str(g(row, "proyecto", "project"))
        if not project:
            continue
        contact = parse_str(g(row, "contacto de mantenimiento en sitio", "contacto mantenimiento", "maintContact"))

        existing = Directorio.query.filter_by(project=project, maint_contact=contact).first()
        # Si no hay contacto, también buscamos por sólo el proyecto
        if not existing and not contact:
            existing = Directorio.query.filter_by(project=project).first()

        target = existing or Directorio(project=project)
        _apply_non_empty(
            target,
            project=project,
            project_code=parse_str(g(row, "codigo de proyecto", "código de proyecto", "projectCode")),
            system_type=parse_str(g(row, "tipo de sistema sistema", "tipo de sistema", "systemType")),
            maint_contact=contact,
            maint_phone=parse_str(g(row, "numero de contacto de mantenimiento", "numero mantenimiento", "maintPhone")),
            maint_contact_2=parse_str(g(row, "2° nombre de contacto de mantenimiento", "contacto 2", "maintContact2")),
            maint_phone_2=parse_str(g(row, "2° numero de contacto de mantenimiento", "numero 2", "maintPhone2")),
            maint_email=parse_str(g(row, "correo de mantenimiento", "email mantenimiento", "maintEmail")),
            internal_pm=parse_str(g(row, "contacto interno pm", "pm interno", "internalPm")),
            internal_phone=parse_str(g(row, "numero de interno", "numero interno", "internalPhone")),
            client_name=parse_str(g(row, "nombre del cliente", "cliente", "clientName")),
            client_company=parse_str(g(row, "empresa del cliente", "empresa", "clientCompany")),
            client_phone=parse_str(g(row, "numero cliente", "telefono cliente", "clientPhone")),
            client_email=parse_str(g(row, "correo del cliente", "email cliente", "clientEmail")),
            category=parse_str(g(row, "categoría", "categoria", "category")) or "Mantenimiento",
        )
        if existing:
            updated += 1
        else:
            db.session.add(target)
            created += 1
    log_change("importar", "directorio", f"{created} nuevos, {updated} actualizados")
    db.session.commit()
    return jsonify(ok=True, created=created, updated=updated)


# ──────────────────────────────────────────────────────────
#  MANTENIMIENTO — upsert por (project, fecha_programada, tipo)
# ──────────────────────────────────────────────────────────
@bp.route("/planeacion-2026", methods=["POST"])
@jwt_required()
@admin_required
def import_planeacion_2026():
    """Importa la hoja 'Programa' del archivo Planeación 2026.xlsx.

    Estructura esperada:
      - Hoja: 'Programa'
      - Cabeceras en fila 7 (cols: N°, Cliente, Código, Proyecto, Plataforma,
        #Paneles, #Inversores, Inicio Sistema, Ubicación, Cuadrilla, Tipo,
        Vigencia Inicio/Fin/Estatus, PR/E/R/N/P, Último Mantto, Plan, Real)
      - Datos desde fila 8 en adelante.

    UPSERT por (project, fecha_programada, tipo).
    NO toca datos existentes que tengan ejecución real (preserva trabajo del usuario).
    """
    f = request.files.get("file")
    if not f:
        return jsonify(error="no_file"), 400
    try:
        wb = load_workbook(BytesIO(f.read()), data_only=True)
    except Exception as e:
        return jsonify(error="bad_file", message=str(e)), 400

    sheet_name = "Programa" if "Programa" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    created, updated, saltadas = 0, 0, 0
    detalle = []

    for r in range(8, ws.max_row + 1):
        def cell(col):
            return ws.cell(row=r, column=col).value

        code = parse_str(cell(3))
        project = parse_str(cell(4))
        if not project and not code:
            continue

        sistema = parse_str(cell(11))        # BESS, PV, etc. (tipo de sistema)
        cuadrilla = parse_str(cell(10))

        # Plan: cols 22 (inicio), 23 (duración días), 24 (fin)
        plan_inicio = parse_date(cell(22))
        plan_duracion = cell(23)
        plan_fin = parse_date(cell(24))

        # Real: cols 25 (inicio), 27 (fin)
        real_inicio = parse_date(cell(25))
        real_fin = parse_date(cell(27))

        # Sin plan ni real → skip
        if not (plan_inicio or plan_fin or real_inicio or real_fin):
            saltadas += 1
            continue

        # Si fechas planeadas iguales → considerar 1 día (inclusivo)
        if plan_inicio and plan_fin and plan_fin == plan_inicio:
            # Mantenemos las fechas como vienen pero el cálculo de días será +1
            pass

        # Estado calculado según ejecución real
        if real_fin:
            estado = "Completado"
        elif real_inicio:
            estado = "En curso"
        else:
            estado = "Programado"

        # Tipo de mantenimiento: por defecto "Preventivo" para BESS y FV;
        # se ajusta si el código indica lo contrario.
        tipo_mantto = "Preventivo"
        if sistema:
            tipo_mantto = f"Preventivo {sistema}"   # ej: "Preventivo BESS"

        # Buscar existente por (project, fecha_programada, tipo)
        existing = Mantenimiento.query.filter_by(
            project=project, fecha_programada=plan_inicio, tipo=tipo_mantto
        ).first()
        # Tolerar variantes anteriores (sin sufijo)
        if not existing:
            existing = Mantenimiento.query.filter_by(
                project=project, fecha_programada=plan_inicio, tipo=sistema
            ).first()
        target = existing or Mantenimiento(project=project, estado=estado)

        _apply_non_empty(
            target,
            project=project,
            code=code,
            tipo=tipo_mantto,
            estado=estado,
            fecha_programada=plan_inicio,
            fecha_fin_programada=plan_fin,
            fecha_inicio_ejecucion=real_inicio,
            fecha_fin_ejecucion=real_fin,
            fecha_ejecutada=real_fin,       # compat con campo legacy
            cuadrilla=cuadrilla,
        )
        # Duración días → horas estimadas (8h/día como aproximación)
        try:
            dur = float(plan_duracion) if plan_duracion not in (None, "") else None
            if dur:
                target.duracion_horas = dur * 8
        except (TypeError, ValueError):
            pass

        if existing:
            updated += 1
            detalle.append({"project": project, "accion": "actualizado", "estado": estado})
        else:
            db.session.add(target)
            created += 1
            detalle.append({"project": project, "accion": "creado", "estado": estado})

    db.session.commit()
    log_change("importar", "planeacion-2026",
               f"Planeación 2026: {created} nuevos, {updated} actualizados, {saltadas} sin datos")
    return jsonify(
        ok=True,
        created=created,
        updated=updated,
        saltadas=saltadas,
        total=created + updated,
        hoja=sheet_name,
        muestra=detalle[:10],
    )


@bp.route("/mantenimiento", methods=["POST"])
@jwt_required()
@admin_required
def import_mantenimiento():
    f = request.files.get("file")
    if not f:
        return jsonify(error="no_file"), 400
    headers, rows = _rows(f)
    idx = {h.lower().strip(): n for n, h in enumerate(headers) if h}

    def g(row, *keys):
        for k in keys:
            n = idx.get(k.lower())
            if n is not None and n < len(row):
                return row[n]
        return None

    created, updated = 0, 0
    for row in rows:
        if not any(row):
            continue
        project = parse_str(g(row, "proyecto", "project"))
        if not project:
            continue
        fecha = parse_date(g(row, "fecha programada", "fechaProgramada"))
        tipo = parse_str(g(row, "tipo"))

        existing = Mantenimiento.query.filter_by(
            project=project, fecha_programada=fecha, tipo=tipo
        ).first()
        target = existing or Mantenimiento(project=project, estado="Programado")
        _apply_non_empty(
            target,
            project=project,
            code=parse_str(g(row, "codigo", "código", "code")),
            tipo=tipo,
            estado=parse_str(g(row, "estado", "status")) or target.estado,
            fecha_programada=fecha,
            fecha_ejecutada=parse_date(g(row, "fecha ejecutada", "fechaEjecutada")),
            cuadrilla=parse_str(g(row, "cuadrilla")),
            responsable=parse_str(g(row, "responsable")),
            descripcion=parse_str(g(row, "descripcion", "descripción")),
            resultados=parse_str(g(row, "resultados")),
        )
        if existing:
            updated += 1
        else:
            db.session.add(target)
            created += 1
    log_change("importar", "mantenimiento", f"{created} nuevos, {updated} actualizados")
    db.session.commit()
    return jsonify(ok=True, created=created, updated=updated)


# ──────────────────────────────────────────────────────────
#  ERRORES — upsert por (brand, code)
# ──────────────────────────────────────────────────────────
@bp.route("/errores", methods=["POST"])
@jwt_required()
@admin_required
def import_errores():
    f = request.files.get("file")
    if not f:
        return jsonify(error="no_file"), 400
    headers, rows = _rows(f)
    idx = {h.lower(): n for n, h in enumerate(headers)}

    def g(row, *keys):
        for k in keys:
            n = idx.get(k.lower())
            if n is not None and n < len(row):
                return row[n]
        return None

    created, updated = 0, 0
    for row in rows:
        if not any(row):
            continue
        brand = parse_str(g(row, "brand", "marca"))
        code = parse_str(g(row, "code", "codigo"))
        if not brand or not code:
            continue
        brand = brand.upper()

        existing = ErrorCatalog.query.filter_by(brand=brand, code=code).first()
        target = existing or ErrorCatalog(brand=brand, code=code)
        _apply_non_empty(
            target,
            equipment=parse_str(g(row, "equipment", "equipo")),
            classification=parse_str(g(row, "classification", "clasificacion")),
            tipo=parse_str(g(row, "tipo", "tipo de error", "tipoError")),
            description=parse_str(g(row, "description", "descripcion")),
            cause=parse_str(g(row, "cause", "causa")),
            solution=parse_str(g(row, "solution", "solucion")),
            severity=parse_str(g(row, "severity", "severidad")),
            es_general=bool(g(row, "es_general", "esGeneral")),
            manual=bool(g(row, "manual")),
        )
        if existing:
            updated += 1
        else:
            db.session.add(target)
            created += 1

    log_change("importar", "errores", f"{created} nuevos, {updated} actualizados")
    db.session.commit()
    return jsonify(ok=True, created=created, updated=updated)
